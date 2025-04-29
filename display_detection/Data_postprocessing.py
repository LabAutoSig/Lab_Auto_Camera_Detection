import subprocess
import sys
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series, ScatterChart

def data_processing_excel(date_time_df,ids_df,results_df,combined_dataframe):
    # Current date and time
    current_date_time = datetime.now()
    current_date = current_date_time.date()
    current_time = current_date_time.time()
    ids = ids_df['IDs']
    units = []
    
    # _____________VWR MicroStar 12 centrifuge________________________
    if ids[0] == 4 and ids[1] == 1 or ids[0] == 1 and ids[1] ==4:
        # Path to the Excel output file
        output_path = 'display_detection\vwr_cenrtifuge.xlsx'
        results_df['units'] = ["RPM/RCF*1000","min"]
        title_device = "VWR MicroStar 12 centrifuge"
        axis_device = "RPM/RCF*1000"
    #_____________________Scales KERN FCE 3K1N________________________________
    elif ids[0] == 42 and ids[1] == 161 or ids[0] == 161 and ids[1] ==42:
        # Path to the Excel output file
        output_path = 'display_detection\kern_scale.xlsx'
        results_df['units'] = ["g", ""]
        title_device = "Scales KERN FCE 3K1N"
        axis_device = "g"
    #_____________________Scales CS Series CS200_________________________________
    elif ids[0] == 5 and ids[1] == 8 or ids[0] == 8 and ids[1] == 5:
        # Path to the Excel output file
        output_path = 'display_detection\cs_series_scale.xlsx'
        results_df['units'] = ["g", ""]
        title_device = "Scales CS Series CS200"
        axis_device = "g"
    #_____________________Thermo Scientific Digital Shaking Drybath 88880028__________
    elif ids[0] == 2 and ids[1] == 314 or ids[0] == 314 and ids[1] == 2:
        # Path to the Excel output file
        output_path = 'display_detection\thermo_scientific_drybath.xlsx'
        results_df['units'] = ["°C","°C"]
        title_device = "Thermo Scientific Digital Shaking Drybath 88880028"
        axis_device = "°C"
    #_____________________Phoenix Instrument Heating and magnetic stirrer RSM-10HP ___________
    elif ids[0] == 2 and ids[1] == 1 or ids[0] == 1 and ids[1] == 2:
        # Path to the Excel output file
        output_path = 'display_detection\heating_stirring_phoenix.xlsx'
        results_df['units'] = ["°C", "1/min"]
        title_device = "Phoenix Instrument Heating and magnetic stirrer RSM-10HP"
        axis_device = "°C"
    #_____________________XSInstruments pH 50+ DHS S/N180356077___________
    elif ids[0] == 3 and ids[1] == 4 or ids[0] == 4 and ids[1] == 3:
        # Path to the Excel output file
        output_path = 'display_detection\XSInstruments_pH.xlsx'
        results_df['units'] = ["pH", "°C"]
        title_device= "XSInstruments pH 50+ DHS S/N180356077"
        axis_device = "pH"

    # Load existing workbook or create a new one
    if os.path.exists(output_path):
        workbook = load_workbook(output_path)
    else:
        workbook = Workbook()
        # Remove the default sheet created by Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    # Define the sheet name as the current date (e.g., '2024-04-27')
    sheet_name = str(current_date)
    chart = None
    # Create a new sheet for the current date if it doesn't exist
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        chart = sheet._charts[0]
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    # Find the next empty column in the sheet
    next_empty_col = sheet.max_column + 1
    next_empty_row = sheet.max_row + 1
    # Write date and time in separate cells if new day
    sheet.cell(row=1, column=next_empty_col, value="Date")
    sheet.cell(row=2, column=next_empty_col, value=current_date)
    sheet.cell(row=1, column=next_empty_col + 1, value="Time")
    sheet.cell(row=2, column=next_empty_col + 1, value=current_time)

    # Function to add dataframe horizontally
    def append_dataframe_horizontal(df, start_col):
        for i, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for j, value in enumerate(row):
                sheet.cell(row=i + 3, column=start_col + j, value=value)

    # Append your dataframes starting from the next empty column
    append_dataframe_horizontal(ids_df, next_empty_col + 2)
    append_dataframe_horizontal(results_df, next_empty_col + 2 + len(ids_df.columns))
    append_dataframe_horizontal(combined_dataframe, next_empty_col + 2 + len(ids_df.columns) + len(results_df.columns))

    # Create or update a scatter chart
    if chart == None:
        print("Chart does not exist")  # If chart doesn't exist, create a new one
        chart = ScatterChart()
        chart.x_axis.title = "Time"
        chart.title = title_device
        chart.y_axis.title = axis_device
        chart.style = 13
        sheet.add_chart(chart, f"A{sheet.max_row + 5}")

    # Use the existing time column (no need to write it again)
    time_col_index = next_empty_col + 1  # This assumes the time column is right next to the date column
    last_time_row = next_empty_row + 1  # Last row where the time value will be written

    # Define the data ranges for the chart
    time_range = Reference(sheet, min_col=time_col_index, min_row=2, max_row=2)
    predicted_values = Reference(sheet, min_col=next_empty_col + 3, min_row=4, max_row=4)

    # Create a series for the chart
    series = Series(predicted_values, xvalues=time_range, title="Predicted Value")
    # Set marker properties (assuming default visibility settings are fine)
    series.marker.symbol = 'circle'  # You can choose other symbols like 'square', 'diamond', etc.
    series.marker.size = 5  # Adjust size to make points visible
    series.marker.solidFill = "FF0000"  # Set marker color to red

    # Clear previous series and append the new one
    if chart is not None:  # Ensure that chart is not None
        #chart.series.clear()  # Clear existing series to avoid duplicating data
        chart.series.append(series)  # Append the new series to the chart

    # Save the workbook
    workbook.save(output_path)
    return
